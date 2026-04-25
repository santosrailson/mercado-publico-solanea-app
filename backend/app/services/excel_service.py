from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from io import BytesIO
from datetime import datetime
from typing import List

from app.models.models import Cessionario, Pagamento, Situacao


def generate_cessionarios_excel(cessionarios: List[Cessionario]) -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = "Cessionários"
    
    # Header
    headers = ['Nome', 'Box/Ponto', 'Atividade', 'Telefone', 'Situação', 'Valor Ref.', 'Periodicidade']
    ws.append(headers)
    
    # Estilo do header
    header_fill = PatternFill(start_color="00A87A", end_color="00A87A", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")
    
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center')
    
    # Dados
    for c in cessionarios:
        ws.append([
            c.nome,
            c.numero_box,
            c.atividade,
            c.telefone,
            c.situacao.value,
            c.valor_referencia,
            c.periodicidade_referencia.value if c.periodicidade_referencia else ''
        ])
    
    # Ajustar largura das colunas
    for idx, col in enumerate(ws.columns, 1):
        max_length = 0
        column = get_column_letter(idx)
        for cell in col:
            try:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
            except:
                pass
        adjusted_width = min(max_length + 2, 50)
        ws.column_dimensions[column].width = adjusted_width
    
    # Congelar primeira linha
    ws.freeze_panes = 'A2'
    
    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer.getvalue()


def generate_pagamentos_excel(pagamentos: List[Pagamento]) -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = "Pagamentos"
    
    headers = ['Cessionário', 'Box/Ponto', 'Data Pagamento', 'Periodicidade', 'Referência Mês', 'Valor', 'Observações']
    ws.append(headers)
    
    # Estilo do header
    header_fill = PatternFill(start_color="00A87A", end_color="00A87A", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")
    
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center')
    
    total = 0.0
    for p in pagamentos:
        ws.append([
            p.cessionario.nome if p.cessionario else '-',
            p.cessionario.numero_box if p.cessionario else '-',
            p.data_pagamento.strftime('%d/%m/%Y'),
            p.periodicidade.value,
            p.referencia_mes,
            p.valor,
            p.observacoes
        ])
        total += p.valor
    
    # Linha de total
    ws.append([''])
    total_row = ['TOTAL', '', '', '', '', total, '']
    ws.append(total_row)
    
    # Estilo da linha total
    for cell in ws[ws.max_row]:
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color="E8F5E9", end_color="E8F5E9", fill_type="solid")
    
    # Ajustar largura
    for idx, col in enumerate(ws.columns, 1):
        max_length = 0
        column = get_column_letter(idx)
        for cell in col:
            try:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
            except:
                pass
        adjusted_width = min(max_length + 2, 50)
        ws.column_dimensions[column].width = adjusted_width
    
    # Formato de moeda para coluna de valor (F)
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row-2, min_col=6, max_col=6):
        for cell in row:
            cell.number_format = 'R$ #,##0.00'
    
    ws.freeze_panes = 'A2'
    
    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer.getvalue()


def generate_cobranca_excel(cessionarios: List[Cessionario]) -> bytes:
    """Gera planilha de cobrança com último pagamento de cada cessionário"""
    wb = Workbook()
    ws = wb.active
    ws.title = "Cobrança"
    
    headers = ['Nome', 'Box/Ponto', 'Último Pagamento', 'Valor Pago', 'Periodicidade', 'Situação']
    ws.append(headers)
    
    header_fill = PatternFill(start_color="00A87A", end_color="00A87A", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")
    
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
    
    for c in cessionarios:
        ultimo_pag = None
        if c.pagamentos:
            ultimo_pag = sorted(c.pagamentos, key=lambda x: x.data_pagamento, reverse=True)[0]
        
        ws.append([
            c.nome,
            c.numero_box,
            ultimo_pag.data_pagamento.strftime('%d/%m/%Y') if ultimo_pag else 'N/A',
            ultimo_pag.valor if ultimo_pag else c.valor_referencia,
            ultimo_pag.periodicidade.value if ultimo_pag else (c.periodicidade_referencia.value if c.periodicidade_referencia else 'Mensal'),
            c.situacao.value
        ])
    
    # Ajustar largura
    for idx, col in enumerate(ws.columns, 1):
        max_length = 0
        column = get_column_letter(idx)
        for cell in col:
            try:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
            except:
                pass
        ws.column_dimensions[column].width = min(max_length + 2, 50)
    
    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer.getvalue()
