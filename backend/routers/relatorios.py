import io
from datetime import datetime, date, timedelta
from typing import Optional, List

from fastapi import APIRouter, Depends, HTTPException, Query
from fastapi.responses import Response, StreamingResponse

from database import get_db
from models import DashStats, GraficoMes, TopPagador, AtividadeRecente, CertidaoResponse
from auth import get_current_user, require_admin
from routers.cessionarios import detectar_ausencias
from collections import Counter

router = APIRouter(prefix="/relatorios", tags=["relatorios"])

# ── Helpers ─────────────────────────────────────────────────────────────────

HEADER_COLOR_HEX = (0x1E, 0x3A, 0x5F)  # #1e3a5f
ACCENT_COLOR_HEX = (0x25, 0x63, 0xEB)  # #2563eb
SUCCESS_COLOR = (0x16, 0xA3, 0x4A)
DANGER_COLOR = (0xDC, 0x26, 0x26)


def _build_pdf_cessionarios(rows, title="Relatório de Cessionários") -> bytes:
    from reportlab.lib import colors
    from reportlab.lib.pagesizes import A4, landscape
    from reportlab.lib.units import cm
    from reportlab.platypus import (
        SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, HRFlowable
    )
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib.enums import TA_CENTER, TA_LEFT

    buf = io.BytesIO()
    doc = SimpleDocTemplate(
        buf,
        pagesize=landscape(A4),
        leftMargin=1.5 * cm,
        rightMargin=1.5 * cm,
        topMargin=2 * cm,
        bottomMargin=2 * cm,
    )
    styles = getSampleStyleSheet()
    primary = colors.HexColor("#1e3a5f")
    accent = colors.HexColor("#2563eb")
    success = colors.HexColor("#16a34a")
    danger = colors.HexColor("#dc2626")

    elements = []

    # Header
    header_style = ParagraphStyle(
        "Header", parent=styles["Heading1"],
        fontSize=16, textColor=primary, alignment=TA_CENTER
    )
    sub_style = ParagraphStyle(
        "Sub", parent=styles["Normal"],
        fontSize=9, textColor=colors.HexColor("#64748b"), alignment=TA_CENTER
    )
    elements.append(Paragraph("PREFEITURA MUNICIPAL DE SOLÂNEA – PB", header_style))
    elements.append(Paragraph("Mercado Público Municipal", sub_style))
    elements.append(Spacer(1, 0.2 * cm))
    elements.append(HRFlowable(width="100%", thickness=2, color=primary))
    elements.append(Spacer(1, 0.3 * cm))
    elements.append(Paragraph(title, ParagraphStyle(
        "T", parent=styles["Heading2"], fontSize=13, textColor=accent, alignment=TA_CENTER
    )))
    elements.append(Paragraph(
        f"Emitido em: {datetime.now().strftime('%d/%m/%Y às %H:%M')}",
        sub_style
    ))
    elements.append(Spacer(1, 0.5 * cm))

    # Table
    headers = ["Box", "Nome", "Atividade", "Telefone", "Situação", "Valor Ref.", "Periodicidade", "Total Pago", "Último Pag."]
    data = [headers]
    for r in rows:
        sit_color = success if r.get("situacao") == "Regular" else danger
        data.append([
            r.get("numero_box") or "-",
            r.get("nome") or "-",
            r.get("atividade") or "-",
            r.get("telefone") or "-",
            r.get("situacao") or "-",
            f"R$ {r.get('valor_ref', 0):.2f}",
            r.get("per_ref") or "-",
            f"R$ {r.get('total_pago', 0):.2f}",
            r.get("ultimo_pagamento") or "-",
        ])

    col_widths = [2 * cm, 5.5 * cm, 3.5 * cm, 3.5 * cm, 2.5 * cm, 2.5 * cm, 3 * cm, 2.5 * cm, 3 * cm]
    table = Table(data, colWidths=col_widths, repeatRows=1)
    table_style = TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), primary),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
        ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
        ("FONTSIZE", (0, 0), (-1, 0), 9),
        ("ALIGN", (0, 0), (-1, -1), "CENTER"),
        ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
        ("FONTSIZE", (0, 1), (-1, -1), 8),
        ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#f1f5f9")]),
        ("GRID", (0, 0), (-1, -1), 0.3, colors.HexColor("#e2e8f0")),
        ("TOPPADDING", (0, 0), (-1, -1), 5),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 5),
    ])
    # Color situação column per value
    for i, r in enumerate(rows):
        row_idx = i + 1
        sit = r.get("situacao", "Regular")
        c = success if sit == "Regular" else danger
        table_style.add("TEXTCOLOR", (4, row_idx), (4, row_idx), c)
        table_style.add("FONTNAME", (4, row_idx), (4, row_idx), "Helvetica-Bold")
    table.setStyle(table_style)
    elements.append(table)

    # Footer
    elements.append(Spacer(1, 0.5 * cm))
    elements.append(HRFlowable(width="100%", thickness=0.5, color=colors.HexColor("#e2e8f0")))
    elements.append(Paragraph(
        "Prefeitura Municipal de Solânea – PB | Mercado Público Municipal | Documento gerado automaticamente pelo Sistema de Gestão",
        ParagraphStyle("Footer", parent=styles["Normal"], fontSize=7,
                       textColor=colors.HexColor("#94a3b8"), alignment=TA_CENTER)
    ))

    doc.build(elements)
    return buf.getvalue()


def _build_pdf_pagamentos(rows, title="Relatório de Pagamentos") -> bytes:
    from reportlab.lib import colors
    from reportlab.lib.pagesizes import A4, landscape
    from reportlab.lib.units import cm
    from reportlab.platypus import (
        SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, HRFlowable
    )
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib.enums import TA_CENTER

    buf = io.BytesIO()
    doc = SimpleDocTemplate(
        buf, pagesize=landscape(A4),
        leftMargin=1.5 * cm, rightMargin=1.5 * cm,
        topMargin=2 * cm, bottomMargin=2 * cm,
    )
    styles = getSampleStyleSheet()
    primary = colors.HexColor("#1e3a5f")
    accent = colors.HexColor("#2563eb")

    elements = []
    h1 = ParagraphStyle("H1", parent=styles["Heading1"], fontSize=16, textColor=primary, alignment=TA_CENTER)
    h2 = ParagraphStyle("H2", parent=styles["Heading2"], fontSize=13, textColor=accent, alignment=TA_CENTER)
    sub = ParagraphStyle("Sub", parent=styles["Normal"], fontSize=9, textColor=colors.HexColor("#64748b"), alignment=TA_CENTER)

    elements.append(Paragraph("PREFEITURA MUNICIPAL DE SOLÂNEA – PB", h1))
    elements.append(Paragraph("Mercado Público Municipal", sub))
    elements.append(Spacer(1, 0.2 * cm))
    elements.append(HRFlowable(width="100%", thickness=2, color=primary))
    elements.append(Spacer(1, 0.3 * cm))
    elements.append(Paragraph(title, h2))
    elements.append(Paragraph(f"Emitido em: {datetime.now().strftime('%d/%m/%Y às %H:%M')}", sub))
    elements.append(Spacer(1, 0.5 * cm))

    total = sum(r.get("valor", 0) for r in rows)
    headers = ["Data", "Cessionário", "Box", "Periodicidade", "Valor", "Cadastrador", "Observação"]
    data = [headers]
    for r in rows:
        data.append([
            r.get("data") or "-",
            r.get("cessionario_nome") or "-",
            r.get("cessionario_box") or "-",
            r.get("periodicidade") or "-",
            f"R$ {r.get('valor', 0):.2f}",
            r.get("usuario_nome") or "-",
            (r.get("observacao") or "")[:40],
        ])
    data.append(["", "", "", "TOTAL", f"R$ {total:.2f}", "", ""])

    col_widths = [2.5 * cm, 6 * cm, 2 * cm, 3 * cm, 2.5 * cm, 4 * cm, 5 * cm]
    table = Table(data, colWidths=col_widths, repeatRows=1)
    table.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), primary),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
        ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
        ("FONTSIZE", (0, 0), (-1, 0), 9),
        ("ALIGN", (0, 0), (-1, -1), "CENTER"),
        ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
        ("FONTSIZE", (0, 1), (-1, -1), 8),
        ("ROWBACKGROUNDS", (0, 1), (-1, -2), [colors.white, colors.HexColor("#f1f5f9")]),
        ("GRID", (0, 0), (-1, -1), 0.3, colors.HexColor("#e2e8f0")),
        ("TOPPADDING", (0, 0), (-1, -1), 5),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 5),
        ("BACKGROUND", (0, -1), (-1, -1), colors.HexColor("#1e3a5f")),
        ("TEXTCOLOR", (0, -1), (-1, -1), colors.white),
        ("FONTNAME", (0, -1), (-1, -1), "Helvetica-Bold"),
    ]))
    elements.append(table)
    elements.append(Spacer(1, 0.5 * cm))
    elements.append(HRFlowable(width="100%", thickness=0.5, color=colors.HexColor("#e2e8f0")))
    elements.append(Paragraph(
        "Prefeitura Municipal de Solânea – PB | Mercado Público Municipal | Documento gerado automaticamente pelo Sistema de Gestão",
        ParagraphStyle("Footer", parent=styles["Normal"], fontSize=7,
                       textColor=colors.HexColor("#94a3b8"), alignment=TA_CENTER)
    ))
    doc.build(elements)
    return buf.getvalue()


def _build_pdf_certidao(cert: dict) -> bytes:
    from reportlab.lib import colors
    from reportlab.lib.pagesizes import A4
    from reportlab.lib.units import cm
    from reportlab.platypus import (
        SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, HRFlowable
    )
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib.enums import TA_CENTER, TA_LEFT, TA_JUSTIFY

    buf = io.BytesIO()
    doc = SimpleDocTemplate(
        buf, pagesize=A4,
        leftMargin=2.5 * cm, rightMargin=2.5 * cm,
        topMargin=2.5 * cm, bottomMargin=2.5 * cm,
    )
    styles = getSampleStyleSheet()
    primary = colors.HexColor("#1e3a5f")
    accent = colors.HexColor("#2563eb")
    success = colors.HexColor("#16a34a")
    danger = colors.HexColor("#dc2626")

    sit_color = success if cert.get("situacao") == "Regular" else danger

    elements = []
    h1 = ParagraphStyle("H1", parent=styles["Heading1"], fontSize=14, textColor=primary, alignment=TA_CENTER)
    sub = ParagraphStyle("Sub", parent=styles["Normal"], fontSize=9, textColor=colors.HexColor("#64748b"), alignment=TA_CENTER)
    body = ParagraphStyle("Body", parent=styles["Normal"], fontSize=10, textColor=colors.HexColor("#0f172a"), alignment=TA_JUSTIFY, leading=16)
    label = ParagraphStyle("Label", parent=styles["Normal"], fontSize=9, textColor=colors.HexColor("#64748b"))
    value = ParagraphStyle("Value", parent=styles["Normal"], fontSize=10, textColor=primary)

    elements.append(Paragraph("PREFEITURA MUNICIPAL DE SOLÂNEA – PB", h1))
    elements.append(Paragraph("Secretaria Municipal de Administração", sub))
    elements.append(Paragraph("Mercado Público Municipal", sub))
    elements.append(Spacer(1, 0.3 * cm))
    elements.append(HRFlowable(width="100%", thickness=3, color=primary))
    elements.append(Spacer(1, 0.5 * cm))

    elements.append(Paragraph(
        "CERTIDÃO DE SITUAÇÃO CADASTRAL",
        ParagraphStyle("Cert", parent=styles["Heading1"], fontSize=16, textColor=primary, alignment=TA_CENTER, spaceAfter=4)
    ))
    elements.append(Paragraph(f"Nº {cert.get('numero', '')}", ParagraphStyle(
        "Num", parent=styles["Normal"], fontSize=10, textColor=accent, alignment=TA_CENTER
    )))
    elements.append(Spacer(1, 0.5 * cm))

    # Info table
    sit = cert.get("situacao", "Irregular")
    info_data = [
        [Paragraph("Cessionário:", label), Paragraph(cert.get("cessionario", "-"), value)],
        [Paragraph("Box:", label), Paragraph(cert.get("numero_box") or "-", value)],
        [Paragraph("Atividade:", label), Paragraph(cert.get("atividade") or "-", value)],
        [Paragraph("Situação:", label), Paragraph(
            sit,
            ParagraphStyle("Sit", parent=styles["Normal"], fontSize=11, textColor=sit_color, fontName="Helvetica-Bold")
        )],
        [Paragraph("Periodicidade:", label), Paragraph(cert.get("periodicidade", "-"), value)],
        [Paragraph("Último Pagamento:", label), Paragraph(cert.get("ultimo_pagamento") or "Nenhum", value)],
        [Paragraph("Total Pago:", label), Paragraph(f"R$ {cert.get('total_pago', 0):.2f}", value)],
        [Paragraph("Valor de Referência:", label), Paragraph(f"R$ {cert.get('valor_ref', 0):.2f}", value)],
    ]
    info_table = Table(info_data, colWidths=[5 * cm, 11 * cm])
    info_table.setStyle(TableStyle([
        ("VALIGN", (0, 0), (-1, -1), "TOP"),
        ("TOPPADDING", (0, 0), (-1, -1), 6),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 6),
        ("LINEBELOW", (0, 0), (-1, -2), 0.3, colors.HexColor("#e2e8f0")),
    ]))
    elements.append(info_table)
    elements.append(Spacer(1, 0.5 * cm))

    # Ausencias
    ausencias = cert.get("ausencias", [])
    if ausencias:
        elements.append(Paragraph(
            f"Períodos sem Pagamento Detectados ({len(ausencias)}):",
            ParagraphStyle("AU", parent=styles["Heading3"], fontSize=10, textColor=danger)
        ))
        aus_text = ", ".join(ausencias[:20])
        if len(ausencias) > 20:
            aus_text += f" ... (+{len(ausencias)-20} mais)"
        elements.append(Paragraph(aus_text, ParagraphStyle(
            "AUL", parent=styles["Normal"], fontSize=9, textColor=danger
        )))
        elements.append(Spacer(1, 0.3 * cm))
    else:
        elements.append(Paragraph(
            "Nenhum período sem pagamento detectado.",
            ParagraphStyle("OK", parent=styles["Normal"], fontSize=10, textColor=success)
        ))
        elements.append(Spacer(1, 0.3 * cm))

    # Legal text
    elements.append(HRFlowable(width="100%", thickness=0.5, color=colors.HexColor("#e2e8f0")))
    elements.append(Spacer(1, 0.3 * cm))
    elements.append(Paragraph(
        f"A presente certidão é emitida para os devidos fins, com base nos registros do Sistema de Gestão "
        f"do Mercado Público Municipal de Solânea – PB, referente à situação cadastral do(a) cessionário(a) "
        f"acima identificado(a), na data de emissão deste documento.",
        body
    ))
    elements.append(Spacer(1, 0.5 * cm))

    emitido_em = cert.get("emitido_em", "")
    try:
        dt = datetime.fromisoformat(emitido_em)
        emitido_fmt = dt.strftime("%d de %B de %Y às %H:%M")
    except Exception:
        emitido_fmt = emitido_em

    elements.append(Paragraph(
        f"Solânea – PB, {emitido_fmt}",
        ParagraphStyle("Local", parent=styles["Normal"], fontSize=9, textColor=colors.HexColor("#64748b"), alignment=TA_CENTER)
    ))
    elements.append(Spacer(1, 0.5 * cm))
    elements.append(Paragraph(
        f"Emitido por: {cert.get('emitido_por', 'Sistema')}",
        ParagraphStyle("EP", parent=styles["Normal"], fontSize=9, textColor=colors.HexColor("#64748b"), alignment=TA_CENTER)
    ))
    elements.append(Spacer(1, 0.8 * cm))
    elements.append(HRFlowable(width="100%", thickness=0.5, color=primary))
    elements.append(Paragraph(
        "Prefeitura Municipal de Solânea – PB | CNPJ: 09.007.245/0001-47 | Mercado Público Municipal | "
        "Documento com validade de 30 dias a partir da data de emissão",
        ParagraphStyle("Foot", parent=styles["Normal"], fontSize=7,
                       textColor=colors.HexColor("#94a3b8"), alignment=TA_CENTER)
    ))
    doc.build(elements)
    return buf.getvalue()


def _build_xlsx_cessionarios(rows) -> bytes:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Cessionários"

    # Header style
    header_fill = PatternFill(start_color="1E3A5F", end_color="1E3A5F", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=10)
    center = Alignment(horizontal="center", vertical="center")
    thin = Border(
        left=Side(style="thin", color="E2E8F0"),
        right=Side(style="thin", color="E2E8F0"),
        top=Side(style="thin", color="E2E8F0"),
        bottom=Side(style="thin", color="E2E8F0"),
    )

    # Title row
    ws.merge_cells("A1:I1")
    ws["A1"] = "MERCADO PÚBLICO MUNICIPAL DE SOLÂNEA – PB"
    ws["A1"].font = Font(bold=True, color="1E3A5F", size=12)
    ws["A1"].alignment = center

    ws.merge_cells("A2:I2")
    ws["A2"] = f"Relatório de Cessionários – Emitido em {datetime.now().strftime('%d/%m/%Y %H:%M')}"
    ws["A2"].font = Font(color="64748B", size=9)
    ws["A2"].alignment = center

    headers = ["Box", "Nome", "Atividade", "Telefone", "Situação", "Valor Ref.", "Periodicidade", "Total Pago", "Último Pag."]
    for col, h in enumerate(headers, start=1):
        cell = ws.cell(row=4, column=col, value=h)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = center
        cell.border = thin

    # Data rows
    alt_fill = PatternFill(start_color="F1F5F9", end_color="F1F5F9", fill_type="solid")
    reg_font = Font(color="16A34A", bold=True)
    irr_font = Font(color="DC2626", bold=True)

    for i, r in enumerate(rows):
        row_num = i + 5
        fill = alt_fill if i % 2 == 1 else None
        values = [
            r.get("numero_box") or "",
            r.get("nome") or "",
            r.get("atividade") or "",
            r.get("telefone") or "",
            r.get("situacao") or "",
            r.get("valor_ref", 0),
            r.get("per_ref") or "",
            r.get("total_pago", 0),
            r.get("ultimo_pagamento") or "",
        ]
        for col, v in enumerate(values, start=1):
            cell = ws.cell(row=row_num, column=col, value=v)
            cell.alignment = center
            cell.border = thin
            if fill:
                cell.fill = fill
            if col == 5:
                cell.font = reg_font if v == "Regular" else irr_font

    # Auto column width
    col_widths = [8, 30, 20, 16, 12, 12, 14, 12, 14]
    for i, w in enumerate(col_widths, start=1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = w

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def _build_xlsx_pagamentos(rows) -> bytes:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Pagamentos"

    header_fill = PatternFill(start_color="1E3A5F", end_color="1E3A5F", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=10)
    center = Alignment(horizontal="center", vertical="center")
    thin = Border(
        left=Side(style="thin", color="E2E8F0"),
        right=Side(style="thin", color="E2E8F0"),
        top=Side(style="thin", color="E2E8F0"),
        bottom=Side(style="thin", color="E2E8F0"),
    )

    ws.merge_cells("A1:G1")
    ws["A1"] = "MERCADO PÚBLICO MUNICIPAL DE SOLÂNEA – PB"
    ws["A1"].font = Font(bold=True, color="1E3A5F", size=12)
    ws["A1"].alignment = center
    ws.merge_cells("A2:G2")
    ws["A2"] = f"Relatório de Pagamentos – Emitido em {datetime.now().strftime('%d/%m/%Y %H:%M')}"
    ws["A2"].font = Font(color="64748B", size=9)
    ws["A2"].alignment = center

    headers = ["Data", "Cessionário", "Box", "Periodicidade", "Valor", "Cadastrador", "Observação"]
    for col, h in enumerate(headers, start=1):
        cell = ws.cell(row=4, column=col, value=h)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = center
        cell.border = thin

    alt_fill = PatternFill(start_color="F1F5F9", end_color="F1F5F9", fill_type="solid")
    total = 0.0
    for i, r in enumerate(rows):
        row_num = i + 5
        fill = alt_fill if i % 2 == 1 else None
        valor = r.get("valor", 0)
        total += valor
        values = [
            r.get("data") or "",
            r.get("cessionario_nome") or "",
            r.get("cessionario_box") or "",
            r.get("periodicidade") or "",
            valor,
            r.get("usuario_nome") or "",
            r.get("observacao") or "",
        ]
        for col, v in enumerate(values, start=1):
            cell = ws.cell(row=row_num, column=col, value=v)
            cell.alignment = center
            cell.border = thin
            if fill:
                cell.fill = fill

    # Total row
    tr = len(rows) + 5
    total_fill = PatternFill(start_color="1E3A5F", end_color="1E3A5F", fill_type="solid")
    total_font = Font(bold=True, color="FFFFFF")
    ws.cell(row=tr, column=4, value="TOTAL").fill = total_fill
    ws.cell(row=tr, column=4).font = total_font
    ws.cell(row=tr, column=4).alignment = center
    ws.cell(row=tr, column=5, value=total).fill = total_fill
    ws.cell(row=tr, column=5).font = total_font
    ws.cell(row=tr, column=5).alignment = center

    col_widths = [14, 30, 8, 14, 12, 20, 30]
    for i, w in enumerate(col_widths, start=1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = w

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def _build_xlsx_geral(cess_rows, pag_rows, stats) -> bytes:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

    wb = openpyxl.Workbook()

    # Sheet: Resumo
    ws_r = wb.active
    ws_r.title = "Resumo"
    header_fill = PatternFill(start_color="1E3A5F", end_color="1E3A5F", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=11)
    center = Alignment(horizontal="center", vertical="center")

    ws_r.merge_cells("A1:B1")
    ws_r["A1"] = "RELATÓRIO GERAL – MERCADO PÚBLICO DE SOLÂNEA"
    ws_r["A1"].font = Font(bold=True, color="1E3A5F", size=13)
    ws_r["A1"].alignment = center
    ws_r.merge_cells("A2:B2")
    ws_r["A2"] = f"Emitido em {datetime.now().strftime('%d/%m/%Y %H:%M')}"
    ws_r["A2"].font = Font(color="64748B", size=9)
    ws_r["A2"].alignment = center

    kpis = [
        ("Total de Cessionários", stats.get("total_cess", 0)),
        ("Regulares", stats.get("regulares", 0)),
        ("Irregulares", stats.get("irregulares", 0)),
        ("Total Arrecadado", f"R$ {stats.get('total_arrecadado', 0):.2f}"),
        ("Total de Pagamentos", stats.get("total_pagamentos", 0)),
        ("Arrecadação do Mês", f"R$ {stats.get('pagamentos_mensais', 0):.2f}"),
    ]
    for i, (k, v) in enumerate(kpis):
        ws_r.cell(row=4 + i, column=1, value=k).font = Font(bold=True, color="1E3A5F")
        ws_r.cell(row=4 + i, column=2, value=v)
    ws_r.column_dimensions["A"].width = 30
    ws_r.column_dimensions["B"].width = 20

    # Sheet: Cessionários
    ws_c = wb.create_sheet("Cessionários")
    headers_c = ["Box", "Nome", "Atividade", "Situação", "Valor Ref.", "Total Pago"]
    for col, h in enumerate(headers_c, start=1):
        cell = ws_c.cell(row=1, column=col, value=h)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = center
    for i, r in enumerate(cess_rows):
        ws_c.cell(row=i + 2, column=1, value=r.get("numero_box") or "")
        ws_c.cell(row=i + 2, column=2, value=r.get("nome") or "")
        ws_c.cell(row=i + 2, column=3, value=r.get("atividade") or "")
        ws_c.cell(row=i + 2, column=4, value=r.get("situacao") or "")
        ws_c.cell(row=i + 2, column=5, value=r.get("valor_ref", 0))
        ws_c.cell(row=i + 2, column=6, value=r.get("total_pago", 0))
    for i, w in enumerate([8, 30, 20, 12, 12, 12], start=1):
        ws_c.column_dimensions[openpyxl.utils.get_column_letter(i)].width = w

    # Sheet: Pagamentos
    ws_p = wb.create_sheet("Pagamentos")
    headers_p = ["Data", "Cessionário", "Box", "Periodicidade", "Valor", "Cadastrador"]
    for col, h in enumerate(headers_p, start=1):
        cell = ws_p.cell(row=1, column=col, value=h)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = center
    for i, r in enumerate(pag_rows):
        ws_p.cell(row=i + 2, column=1, value=r.get("data") or "")
        ws_p.cell(row=i + 2, column=2, value=r.get("cessionario_nome") or "")
        ws_p.cell(row=i + 2, column=3, value=r.get("cessionario_box") or "")
        ws_p.cell(row=i + 2, column=4, value=r.get("periodicidade") or "")
        ws_p.cell(row=i + 2, column=5, value=r.get("valor", 0))
        ws_p.cell(row=i + 2, column=6, value=r.get("usuario_nome") or "")
    for i, w in enumerate([14, 30, 8, 14, 12, 20], start=1):
        ws_p.column_dimensions[openpyxl.utils.get_column_letter(i)].width = w

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


# ── Dashboard ────────────────────────────────────────────────────────────────

@router.get("/dashboard", response_model=DashStats, tags=["dashboard"])
async def dashboard(user: dict = Depends(get_current_user)):
    async with get_db() as db:
        # Totals
        async with db.execute("SELECT COUNT(*) as c FROM cessionarios") as c:
            total_cess = (await c.fetchone())["c"]
        async with db.execute("SELECT COUNT(*) as c FROM cessionarios WHERE situacao = 'Regular'") as c:
            regulares = (await c.fetchone())["c"]
        irregulares = total_cess - regulares

        async with db.execute("SELECT SUM(valor) as s FROM pagamentos") as c:
            total_arr = (await c.fetchone())["s"] or 0.0

        async with db.execute("SELECT COUNT(*) as c FROM pagamentos") as c:
            total_pags = (await c.fetchone())["c"]

        # This month
        hoje = date.today()
        mes_ini = date(hoje.year, hoje.month, 1).isoformat()
        async with db.execute(
            "SELECT SUM(valor) as s FROM pagamentos WHERE data >= ?", (mes_ini,)
        ) as c:
            pag_mes = (await c.fetchone())["s"] or 0.0

        # Grafico 6 meses
        grafico = []
        for m in range(5, -1, -1):
            ref = hoje.replace(day=1)
            # Go back m months
            year = ref.year
            month = ref.month - m
            while month <= 0:
                month += 12
                year -= 1
            mes_s = date(year, month, 1).isoformat()
            if month == 12:
                prox = date(year + 1, 1, 1).isoformat()
            else:
                prox = date(year, month + 1, 1).isoformat()
            async with db.execute(
                "SELECT SUM(valor) as s FROM pagamentos WHERE data >= ? AND data < ?",
                (mes_s, prox)
            ) as c:
                s = (await c.fetchone())["s"] or 0.0
            nome_mes = date(year, month, 1).strftime("%b/%y")
            grafico.append(GraficoMes(mes=nome_mes, total=s))

        # Top pagadores
        async with db.execute(
            """SELECT c.id, c.nome, SUM(p.valor) as total
               FROM pagamentos p JOIN cessionarios c ON p.cessionario_id = c.id
               GROUP BY c.id ORDER BY total DESC LIMIT 5"""
        ) as c:
            top_rows = await c.fetchall()
        top_pagadores = [TopPagador(id=r["id"], nome=r["nome"], total=r["total"]) for r in top_rows]

        # Atividade recente (últimos 10 pagamentos)
        async with db.execute(
            """SELECT p.criado_em, p.valor, c.nome as cess_nome, u.nome as user_nome
               FROM pagamentos p
               JOIN cessionarios c ON p.cessionario_id = c.id
               LEFT JOIN users u ON p.usuario_id = u.id
               ORDER BY p.criado_em DESC LIMIT 10"""
        ) as c:
            act_rows = await c.fetchall()
        atividade_recente = [
            AtividadeRecente(
                tipo="pagamento",
                descricao=f"Pagamento de R$ {r['valor']:.2f} – {r['cess_nome']}",
                data=r["criado_em"] or "",
                usuario=r["user_nome"] if user.get("is_admin") else None,
            )
            for r in act_rows
        ]

    return DashStats(
        total_cess=total_cess,
        regulares=regulares,
        irregulares=irregulares,
        total_arrecadado=total_arr,
        total_pagamentos=total_pags,
        pagamentos_mensais=pag_mes,
        grafico_6meses=grafico,
        top_pagadores=top_pagadores,
        atividade_recente=atividade_recente,
    )


# ── PDF Endpoints ─────────────────────────────────────────────────────────────

@router.get("/cessionarios/pdf")
async def pdf_cessionarios(
    situacao: str = Query("todos"),
    _: dict = Depends(require_admin),
):
    async with get_db() as db:
        if situacao != "todos":
            async with db.execute(
                "SELECT * FROM cessionarios WHERE situacao = ? ORDER BY nome", (situacao,)
            ) as c:
                rows = await c.fetchall()
        else:
            async with db.execute("SELECT * FROM cessionarios ORDER BY nome") as c:
                rows = await c.fetchall()

        result = []
        for r in rows:
            async with db.execute(
                "SELECT SUM(valor) as total, MAX(data) as ultimo FROM pagamentos WHERE cessionario_id = ?",
                (r["id"],)
            ) as c:
                pr = await c.fetchone()
            result.append({**dict(r), "total_pago": pr["total"] or 0.0, "ultimo_pagamento": pr["ultimo"]})

    title = "Relatório de Cessionários" if situacao == "todos" else f"Cessionários – {situacao}"
    pdf = _build_pdf_cessionarios(result, title)
    return Response(content=pdf, media_type="application/pdf",
                    headers={"Content-Disposition": "attachment; filename=cessionarios.pdf"})


@router.get("/pagamentos/pdf")
async def pdf_pagamentos(
    data_ini: str = Query(""),
    data_fim: str = Query(""),
    _: dict = Depends(require_admin),
):
    async with get_db() as db:
        conditions = []
        params = []
        if data_ini:
            conditions.append("p.data >= ?")
            params.append(data_ini)
        if data_fim:
            conditions.append("p.data <= ?")
            params.append(data_fim)
        where = "WHERE " + " AND ".join(conditions) if conditions else ""
        async with db.execute(
            f"""SELECT p.*, c.nome as cessionario_nome, c.numero_box as cessionario_box,
                       u.nome as usuario_nome
                FROM pagamentos p
                JOIN cessionarios c ON p.cessionario_id = c.id
                LEFT JOIN users u ON p.usuario_id = u.id
                {where} ORDER BY p.data DESC""",
            params
        ) as c:
            rows = await c.fetchall()
    result = [dict(r) for r in rows]
    pdf = _build_pdf_pagamentos(result)
    return Response(content=pdf, media_type="application/pdf",
                    headers={"Content-Disposition": "attachment; filename=pagamentos.pdf"})


@router.get("/geral/pdf")
async def pdf_geral(_: dict = Depends(require_admin)):
    async with get_db() as db:
        async with db.execute("SELECT * FROM cessionarios ORDER BY nome") as c:
            cess_rows = await c.fetchall()
        async with db.execute(
            """SELECT p.*, c.nome as cessionario_nome, c.numero_box as cessionario_box,
                       u.nome as usuario_nome
                FROM pagamentos p
                JOIN cessionarios c ON p.cessionario_id = c.id
                LEFT JOIN users u ON p.usuario_id = u.id
                ORDER BY p.data DESC"""
        ) as c:
            pag_rows = await c.fetchall()

        result_cess = []
        for r in cess_rows:
            async with db.execute(
                "SELECT SUM(valor) as total, MAX(data) as ultimo FROM pagamentos WHERE cessionario_id = ?",
                (r["id"],)
            ) as c:
                pr = await c.fetchone()
            result_cess.append({**dict(r), "total_pago": pr["total"] or 0.0, "ultimo_pagamento": pr["ultimo"]})

    from reportlab.lib import colors
    from reportlab.lib.pagesizes import A4, landscape
    from reportlab.lib.units import cm
    from reportlab.platypus import (
        SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, HRFlowable, PageBreak
    )
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib.enums import TA_CENTER

    buf = io.BytesIO()
    doc = SimpleDocTemplate(buf, pagesize=landscape(A4),
                             leftMargin=1.5 * cm, rightMargin=1.5 * cm,
                             topMargin=2 * cm, bottomMargin=2 * cm)
    styles = getSampleStyleSheet()
    primary = colors.HexColor("#1e3a5f")
    accent = colors.HexColor("#2563eb")

    h1 = ParagraphStyle("H1", parent=styles["Heading1"], fontSize=16, textColor=primary, alignment=TA_CENTER)
    h2 = ParagraphStyle("H2", parent=styles["Heading2"], fontSize=13, textColor=accent, alignment=TA_CENTER)
    sub = ParagraphStyle("Sub", parent=styles["Normal"], fontSize=9, textColor=colors.HexColor("#64748b"), alignment=TA_CENTER)

    elements = []
    elements.append(Paragraph("PREFEITURA MUNICIPAL DE SOLÂNEA – PB", h1))
    elements.append(Paragraph("Mercado Público Municipal – Relatório Consolidado", sub))
    elements.append(Spacer(1, 0.2 * cm))
    elements.append(HRFlowable(width="100%", thickness=2, color=primary))
    elements.append(Spacer(1, 0.3 * cm))
    elements.append(Paragraph(f"Emitido em: {datetime.now().strftime('%d/%m/%Y às %H:%M')}", sub))
    elements.append(Spacer(1, 0.5 * cm))

    total_arr = sum(r.get("total_pago", 0) for r in result_cess)
    regulares = sum(1 for r in result_cess if r.get("situacao") == "Regular")
    kpi_data = [
        ["Total de Cessionários", "Regulares", "Irregulares", "Total Arrecadado"],
        [len(result_cess), regulares, len(result_cess) - regulares, f"R$ {total_arr:.2f}"],
    ]
    kpi_table = Table(kpi_data, colWidths=[6 * cm] * 4)
    kpi_table.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), primary),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
        ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
        ("FONTSIZE", (0, 0), (-1, 0), 10),
        ("FONTSIZE", (0, 1), (-1, 1), 14),
        ("FONTNAME", (0, 1), (-1, 1), "Helvetica-Bold"),
        ("TEXTCOLOR", (0, 1), (-1, 1), primary),
        ("ALIGN", (0, 0), (-1, -1), "CENTER"),
        ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
        ("TOPPADDING", (0, 0), (-1, -1), 8),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 8),
        ("BOX", (0, 0), (-1, -1), 1, primary),
        ("INNERGRID", (0, 0), (-1, -1), 0.3, colors.HexColor("#e2e8f0")),
    ]))
    elements.append(kpi_table)
    elements.append(Spacer(1, 0.8 * cm))

    # ── Cessionários table ──
    elements.append(Paragraph("Cessionários", h2))
    elements.append(Spacer(1, 0.3 * cm))
    success_c = colors.HexColor("#16a34a")
    danger_c  = colors.HexColor("#dc2626")
    cess_headers = ["Box", "Nome", "Atividade", "Situação", "Valor Ref.", "Total Pago", "Último Pag."]
    cess_data = [cess_headers]
    for r in result_cess:
        cess_data.append([
            r.get("numero_box") or "-",
            r.get("nome") or "-",
            r.get("atividade") or "-",
            r.get("situacao") or "-",
            f"R$ {r.get('valor_ref', 0):.2f}",
            f"R$ {r.get('total_pago', 0):.2f}",
            r.get("ultimo_pagamento") or "-",
        ])
    cess_col_widths = [2 * cm, 6 * cm, 4 * cm, 3 * cm, 3 * cm, 3 * cm, 3 * cm]
    cess_table = Table(cess_data, colWidths=cess_col_widths, repeatRows=1)
    cess_ts = TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), primary),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
        ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
        ("FONTSIZE", (0, 0), (-1, -1), 8),
        ("ALIGN", (0, 0), (-1, -1), "CENTER"),
        ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
        ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#f1f5f9")]),
        ("GRID", (0, 0), (-1, -1), 0.3, colors.HexColor("#e2e8f0")),
        ("TOPPADDING", (0, 0), (-1, -1), 4),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 4),
    ])
    for i, r in enumerate(result_cess):
        c_color = success_c if r.get("situacao") == "Regular" else danger_c
        cess_ts.add("TEXTCOLOR", (3, i + 1), (3, i + 1), c_color)
        cess_ts.add("FONTNAME", (3, i + 1), (3, i + 1), "Helvetica-Bold")
    cess_table.setStyle(cess_ts)
    elements.append(cess_table)

    # ── Pagamentos table ──
    pag_dict = [dict(r) for r in pag_rows]
    elements.append(PageBreak())
    elements.append(Paragraph("Pagamentos Registrados", h2))
    elements.append(Spacer(1, 0.3 * cm))
    pag_headers = ["Data", "Cessionário", "Box", "Periodicidade", "Valor", "Cadastrador"]
    pag_data = [pag_headers]
    for r in pag_dict:
        pag_data.append([
            r.get("data") or "-",
            r.get("cessionario_nome") or "-",
            r.get("cessionario_box") or "-",
            r.get("periodicidade") or "-",
            f"R$ {r.get('valor', 0):.2f}",
            r.get("usuario_nome") or "-",
        ])
    pag_col_widths = [3 * cm, 7 * cm, 2.5 * cm, 3.5 * cm, 3 * cm, 5 * cm]
    pag_table = Table(pag_data, colWidths=pag_col_widths, repeatRows=1)
    pag_table.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), primary),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
        ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
        ("FONTSIZE", (0, 0), (-1, -1), 8),
        ("ALIGN", (0, 0), (-1, -1), "CENTER"),
        ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
        ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#f1f5f9")]),
        ("GRID", (0, 0), (-1, -1), 0.3, colors.HexColor("#e2e8f0")),
        ("TOPPADDING", (0, 0), (-1, -1), 4),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 4),
    ]))
    elements.append(pag_table)

    # Footer
    elements.append(Spacer(1, 0.5 * cm))
    elements.append(HRFlowable(width="100%", thickness=0.5, color=colors.HexColor("#e2e8f0")))
    elements.append(Paragraph(
        "Prefeitura Municipal de Solânea – PB | Mercado Público Municipal | Documento gerado automaticamente",
        ParagraphStyle("Foot", parent=styles["Normal"], fontSize=7,
                       textColor=colors.HexColor("#94a3b8"), alignment=TA_CENTER)
    ))

    doc.build(elements)
    return Response(content=buf.getvalue(), media_type="application/pdf",
                    headers={"Content-Disposition": "attachment; filename=relatorio_geral.pdf"})


@router.get("/cobranca/pdf")
async def pdf_cobranca(_: dict = Depends(require_admin)):
    async with get_db() as db:
        async with db.execute("SELECT * FROM cessionarios ORDER BY nome") as c:
            cess_rows = await c.fetchall()
        result = []
        for r in cess_rows:
            async with db.execute(
                "SELECT MAX(data) as ultimo FROM pagamentos WHERE cessionario_id = ?", (r["id"],)
            ) as c:
                pr = await c.fetchone()
            result.append({**dict(r), "ultimo_pagamento": pr["ultimo"]})

    from reportlab.lib import colors
    from reportlab.lib.pagesizes import A4
    from reportlab.lib.units import cm
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, HRFlowable
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib.enums import TA_CENTER

    buf = io.BytesIO()
    doc = SimpleDocTemplate(buf, pagesize=A4,
                             leftMargin=2 * cm, rightMargin=2 * cm,
                             topMargin=2 * cm, bottomMargin=2 * cm)
    styles = getSampleStyleSheet()
    primary = colors.HexColor("#1e3a5f")
    danger = colors.HexColor("#dc2626")

    elements = []
    h1 = ParagraphStyle("H1", parent=styles["Heading1"], fontSize=14, textColor=primary, alignment=TA_CENTER)
    sub = ParagraphStyle("Sub", parent=styles["Normal"], fontSize=9, textColor=colors.HexColor("#64748b"), alignment=TA_CENTER)
    elements.append(Paragraph("PREFEITURA MUNICIPAL DE SOLÂNEA – PB", h1))
    elements.append(Paragraph("Mercado Público Municipal – Lista de Cobrança", sub))
    elements.append(Paragraph(f"Emitido em: {datetime.now().strftime('%d/%m/%Y')}", sub))
    elements.append(Spacer(1, 0.3 * cm))
    elements.append(HRFlowable(width="100%", thickness=2, color=primary))
    elements.append(Spacer(1, 0.5 * cm))

    headers = ["Box", "Cessionário", "Atividade", "Telefone", "Valor Ref.", "Último Pag.", "Situação"]
    data = [headers]
    for r in result:
        data.append([
            r.get("numero_box") or "-",
            r.get("nome") or "-",
            r.get("atividade") or "-",
            r.get("telefone") or "-",
            f"R$ {r.get('valor_ref', 0):.2f}",
            r.get("ultimo_pagamento") or "Nunca",
            r.get("situacao") or "-",
        ])

    col_widths = [1.5 * cm, 5 * cm, 3.5 * cm, 3 * cm, 2.5 * cm, 3 * cm, 2.5 * cm]
    table = Table(data, colWidths=col_widths, repeatRows=1)
    success = colors.HexColor("#16a34a")
    ts = TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), primary),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
        ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
        ("FONTSIZE", (0, 0), (-1, 0), 9),
        ("ALIGN", (0, 0), (-1, -1), "CENTER"),
        ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
        ("FONTSIZE", (0, 1), (-1, -1), 8),
        ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#f1f5f9")]),
        ("GRID", (0, 0), (-1, -1), 0.3, colors.HexColor("#e2e8f0")),
        ("TOPPADDING", (0, 0), (-1, -1), 5),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 5),
    ])
    for i, r in enumerate(result):
        sit = r.get("situacao", "Regular")
        c = success if sit == "Regular" else danger
        ts.add("TEXTCOLOR", (6, i + 1), (6, i + 1), c)
        ts.add("FONTNAME", (6, i + 1), (6, i + 1), "Helvetica-Bold")
    table.setStyle(ts)
    elements.append(table)
    elements.append(Spacer(1, 0.5 * cm))
    elements.append(HRFlowable(width="100%", thickness=0.5, color=colors.HexColor("#e2e8f0")))
    elements.append(Paragraph(
        "Este documento é para uso interno da administração do Mercado Público Municipal de Solânea – PB.",
        ParagraphStyle("F", parent=styles["Normal"], fontSize=7, textColor=colors.HexColor("#94a3b8"), alignment=TA_CENTER)
    ))
    doc.build(elements)
    return Response(content=buf.getvalue(), media_type="application/pdf",
                    headers={"Content-Disposition": "attachment; filename=lista_cobranca.pdf"})


@router.get("/cessionarios/{cid}/certidao/pdf")
async def pdf_certidao(cid: str, user: dict = Depends(get_current_user)):
    from routers.cessionarios import certidao as get_certidao
    # Reuse the certidao logic
    async with get_db() as db:
        async with db.execute("SELECT * FROM cessionarios WHERE id = ?", (cid,)) as c:
            cess = await c.fetchone()
        if not cess:
            raise HTTPException(status_code=404, detail="Cessionário não encontrado")
        if not user.get("is_admin") and cess["criado_por"] != user["sub"]:
            raise HTTPException(status_code=403, detail="Sem permissão")
        async with db.execute(
            "SELECT * FROM pagamentos WHERE cessionario_id = ? ORDER BY data", (cid,)
        ) as c:
            pags = await c.fetchall()
        async with db.execute("SELECT nome FROM users WHERE id = ?", (user["sub"],)) as c:
            u = await c.fetchone()
        emitido_por = u["nome"] if u else "Sistema"

    pagamentos_list = [dict(p) for p in pags]
    if pagamentos_list:
        counter = Counter(p["periodicidade"] for p in pagamentos_list)
        periodicidade = counter.most_common(1)[0][0]
    else:
        periodicidade = cess["per_ref"] or "Mensal"

    ausencias = detectar_ausencias(pagamentos_list, periodicidade)
    ultimo = max((p["data"] for p in pagamentos_list), default=None)
    total_pago = sum(p["valor"] for p in pagamentos_list)
    numero = f"CERT-{cid[:8].upper()}-{datetime.utcnow().strftime('%Y%m%d')}"

    cert = {
        "numero": numero,
        "cessionario": cess["nome"],
        "numero_box": cess["numero_box"],
        "atividade": cess["atividade"],
        "situacao": cess["situacao"],
        "ausencias": ausencias,
        "ultimo_pagamento": ultimo,
        "periodicidade": periodicidade,
        "emitido_em": datetime.utcnow().isoformat(),
        "emitido_por": emitido_por,
        "total_pago": total_pago,
        "valor_ref": cess["valor_ref"],
    }
    pdf = _build_pdf_certidao(cert)
    fname = f"certidao_{cess['nome'].replace(' ', '_')}.pdf"
    return Response(content=pdf, media_type="application/pdf",
                    headers={"Content-Disposition": f"attachment; filename={fname}"})


# ── XLSX Endpoints ────────────────────────────────────────────────────────────

@router.get("/cessionarios/xlsx")
async def xlsx_cessionarios(_: dict = Depends(require_admin)):
    async with get_db() as db:
        async with db.execute("SELECT * FROM cessionarios ORDER BY nome") as c:
            rows = await c.fetchall()
        result = []
        for r in rows:
            async with db.execute(
                "SELECT SUM(valor) as total, MAX(data) as ultimo FROM pagamentos WHERE cessionario_id = ?",
                (r["id"],)
            ) as c:
                pr = await c.fetchone()
            result.append({**dict(r), "total_pago": pr["total"] or 0.0, "ultimo_pagamento": pr["ultimo"]})
    xlsx = _build_xlsx_cessionarios(result)
    return Response(content=xlsx,
                    media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    headers={"Content-Disposition": "attachment; filename=cessionarios.xlsx"})


@router.get("/pagamentos/xlsx")
async def xlsx_pagamentos(_: dict = Depends(require_admin)):
    async with get_db() as db:
        async with db.execute(
            """SELECT p.*, c.nome as cessionario_nome, c.numero_box as cessionario_box,
                       u.nome as usuario_nome
                FROM pagamentos p
                JOIN cessionarios c ON p.cessionario_id = c.id
                LEFT JOIN users u ON p.usuario_id = u.id
                ORDER BY p.data DESC"""
        ) as c:
            rows = await c.fetchall()
    result = [dict(r) for r in rows]
    xlsx = _build_xlsx_pagamentos(result)
    return Response(content=xlsx,
                    media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    headers={"Content-Disposition": "attachment; filename=pagamentos.xlsx"})


@router.get("/geral/xlsx")
async def xlsx_geral(_: dict = Depends(require_admin)):
    async with get_db() as db:
        async with db.execute("SELECT * FROM cessionarios ORDER BY nome") as c:
            cess_rows = await c.fetchall()
        async with db.execute(
            """SELECT p.*, c.nome as cessionario_nome, c.numero_box as cessionario_box,
                       u.nome as usuario_nome
                FROM pagamentos p
                JOIN cessionarios c ON p.cessionario_id = c.id
                LEFT JOIN users u ON p.usuario_id = u.id
                ORDER BY p.data DESC"""
        ) as c:
            pag_rows = await c.fetchall()

        result_cess = []
        for r in cess_rows:
            async with db.execute(
                "SELECT SUM(valor) as total, MAX(data) as ultimo FROM pagamentos WHERE cessionario_id = ?",
                (r["id"],)
            ) as c:
                pr = await c.fetchone()
            result_cess.append({**dict(r), "total_pago": pr["total"] or 0.0, "ultimo_pagamento": pr["ultimo"]})

        async with db.execute("SELECT COUNT(*) as c FROM cessionarios") as c:
            total_cess = (await c.fetchone())["c"]
        async with db.execute("SELECT COUNT(*) as c FROM cessionarios WHERE situacao = 'Regular'") as c:
            regulares = (await c.fetchone())["c"]
        async with db.execute("SELECT SUM(valor) as s FROM pagamentos") as c:
            total_arr = (await c.fetchone())["s"] or 0.0
        async with db.execute("SELECT COUNT(*) as c FROM pagamentos") as c:
            total_pags = (await c.fetchone())["c"]
        hoje = date.today()
        mes_ini = date(hoje.year, hoje.month, 1).isoformat()
        async with db.execute("SELECT SUM(valor) as s FROM pagamentos WHERE data >= ?", (mes_ini,)) as c:
            pag_mes = (await c.fetchone())["s"] or 0.0

    stats = {
        "total_cess": total_cess, "regulares": regulares,
        "irregulares": total_cess - regulares,
        "total_arrecadado": total_arr, "total_pagamentos": total_pags,
        "pagamentos_mensais": pag_mes,
    }
    xlsx = _build_xlsx_geral(result_cess, [dict(r) for r in pag_rows], stats)
    return Response(content=xlsx,
                    media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    headers={"Content-Disposition": "attachment; filename=relatorio_geral.xlsx"})
